from flask import Flask, request, jsonify
import openpyxl
import random
import re
import os
import logging
from datetime import datetime

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# –ü—É—Ç—å –∫ Excel-—Ñ–∞–π–ª—É —Å –≤–æ–ø—Ä–æ—Å–∞–º–∏
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(BASE_DIR, "questions.xlsx")

if not os.path.exists(excel_path):
    raise FileNotFoundError(f"–§–∞–π–ª {excel_path} –Ω–µ –Ω–∞–π–¥–µ–Ω!")

# –ó–∞–≥—Ä—É–∑–∫–∞ Excel —Ñ–∞–π–ª–∞
workbook = openpyxl.load_workbook(excel_path)
sheet_names = workbook.sheetnames

# –°–ª–æ–≤–∞—Ä—å —Å ID –∫–∞—Ä—Ç–∏–Ω–æ–∫ –¥–ª—è –ê–ª–∏—Å—ã
ALICE_IMAGE_IDS = {
  "6668": "1030494/57df06289997d0b975b0",
  "6667": "1652229/8c2f3fed1f1627a74e47",
  "6663": "1652229/baa39a8647b4f6a75fee",
  "6645": "13200873/bda9c083cfafae4764c0",
  "6638": "1652229/57167d8c94edef667ecd",
  "6581": "1656841/53b93a4f853f557f78c7",
  "6580": "13200873/7cb497753a3169364075",
  "6579": "1652229/510382cc5b64b80e5e2a",
  "6578": "1652229/ad19d477852d4983a8bb",
  "6577": "1652229/c86ca9b8bf5e9f90ec71",
  "6576": "1652229/cd6ae0622eae6676b4f7",
  "6450": "13200873/693416b28e1a692bdd35",
  "6448": "13200873/da4d8f3fb5560fbe2f83",
  "6447": "1652229/b5480c2992f320bd982c",
  "6445": "13200873/122e41c1471c376077a0",
  "6440": "1652229/7a11c4d2e5baa27cf694",
  "6439": "1652229/1e32128988a4d5cbc291",
  "6437": "13200873/4cff1e97b4a3874a6ab7",
  "6436": "1652229/70693d3d73372d9ed315",
  "6355": "1652229/8ecad1ff6204ef149c36",
  "6350": "13200873/130105b972bfaad7a6a2",
  "6341": "13200873/922fbc92c02908217f8f",
  "6339": "1652229/1adcfbf7b51b37afbf58",
  "6338": "1030494/0bdfef64f202a8434cdc",
  "6337": "13200873/94d7f9b538632030e7f7",
  "6336": "1652229/98ba45f1c700b4981529",
  "6335": "1652229/9625b4a017127d64b9be",
  "6334": "1652229/a2eda4bbaff5e65b906b",
  "6333": "1656841/9dfdb0d6c95992b5a951",
  "6331": "1030494/c99f14f3b415ff591720",
  "6328": "1656841/99c6872db80e0b98eb1c",
  "6327": "13200873/8f1fa4935070d2e61f5b",
  "6324": "1652229/2717d760e2051d3293ea",
  "6323": "13200873/45bdc3f99ff0e436599c",
  "6320": "1652229/04af5f550c859958f4b1",
  "6319": "1030494/15b814b429e95b27e74f",
  "6318": "1030494/cba0ca11424f30985538",
  "6317": "1030494/6bc1a313139bf200d22c",
  "6316": "1652229/bae40a955c89db0b5b0a",
  "6315": "13200873/a5d09861cae566a92117",
  "6312": "13200873/1956a369d6b86b1f7990",
  "6305": "13200873/fb8be027da02189590b4",
  "5040": "1030494/56b2f96bde63186754d3",
  "5039": "1030494/496003edbb33ab1f46d4",
  "5038": "1652229/aa801f79fb187d450782",
  "5037": "13200873/9da8fa5e416c26c272cc",
  "5036": "1030494/6ee58785fda8ff21f072",
  "5035": "1656841/68c3a48e11fae4cbb8dd",
  "5034": "1652229/6a225d83cba52471e617",
  "5033": "1030494/444f450e629883f6f278",
  "5032": "1652229/7db47257775052429ccc",
  "5031": "1652229/2fc3f4181c88b9c5ee34",
  "5030": "13200873/57a408f70e61e762dad4",
  "5029": "1030494/046ea5fe2a53b970e5b5",
  "5028": "1030494/69eceafa5248fb8cf5b7",
  "5027": "13200873/a5014401e7dbc21f48dd",
  "5026": "1652229/1de926c8b92c0e529c66",
  "5025": "1652229/5097a836778bd12184f5",
  "5024": "13200873/e57f7823fd4c657b89bd",
  "5023": "1652229/9c95d15c5deec031d9ab",
  "5022": "1652229/39ced67aacc92afe3d0f",
  "5021": "13200873/d71d3d4b040d953f6e54",
  "5020": "1652229/1fde3265a1fa92f140bb",
  "5019": "1652229/23252fc7e77fd5cabfe2",
  "5018": "13200873/a1f45a89a159e57df4c8",
  "5017": "13200873/3e394be0f661936986cd",
  "5016": "1652229/7cadae30862e5340386b",
  "5015": "1652229/bc50411764b9bad5aecc",
  "5014": "1030494/324a594712804beb1118",
  "5013": "13200873/8e4acb9daa710ea0ba5b",
  "5012": "1656841/424b1cc5cc3335e30633",
  "5011": "1030494/d250350f5e4da3ebf626",
  "5010": "13200873/6a01f2b046648df608d1",
  "5009": "1652229/60c39748f70d24114be0",
  "5008": "1652229/cbdbc728881f3188a674",
  "5007": "13200873/0c04ee5a29a10467409c",
  "5006": "1652229/3fa8d0501900dfc082bf",
  "5005": "1652229/fd3cd69ac7e41fc3a9cd",
  "5004": "13200873/e91ea27b6054533e4222",
  "5003": "1652229/29030b2d6211b28f4a8b",
  "5002": "13200873/2509fabbce9db3508444",
  "5001": "1652229/0588b20bf0145425d0cc",
  "4443": "1652229/7cd44a8537b52ca29719",
  "4441": "13200873/961049698548660c1edf",
  "2004": "1652229/bbb69d7330833e010ffc",
  "2001": "1652229/5ea053d30ac111cd4234",
  "1414": "13200873/337b2074f50590d411f3",
  "1413": "1030494/5ef86fa45ec54120e658",
  "1412": "1652229/fb02aae9dea10177a903",
  "1411": "1533899/87025db269e38a1082cd",
  "1410": "1652229/90531f188d67d353cf2d",
  "1409": "1652229/b5a9fd4aa93b935bdbba",
  "1408": "13200873/763b268abc5fc756ad2f",
  "1407": "13200873/65c3519909a8e2fbf86a",
  "1406": "1652229/70db0df5a8d8ee886664",
  "1405": "13200873/8f29be55086af4bc6999",
  "1404": "13200873/16e709ddb2a3ef1c0ad8",
  "1403": "13200873/5429645baaa05006b298",
  "1402": "1030494/36357db8ab33d200616c",
  "1401": "1652229/6f94e20db6dbdf7be122",
  "1202": "1652229/d69fdcf9558e6487d63c",
  "1201": "1652229/7245445e8d0d5ff8c31c",
  "1118": "13200873/91f492e641595f44bd23",
  "1117": "13200873/545b0087fe5c12cd5155",
  "1116": "1656841/71344d574c57fa13555f",
  "1115": "1533899/d7025c1a62fef9b1fd71",
  "1114": "13200873/ba8cce54d7f3d4ad71d0",
  "1113": "13200873/42ec11d4f472ce091bda",
  "1112": "1652229/bc11321dd775f5606aa2",
  "1111": "1652229/265d7e92d1185f55adf2"
}


def parse_options(options_str):
    """–ü–∞—Ä—Å–∏–Ω–≥ —Å—Ç—Ä–æ–∫–∏ —Å –≤–∞—Ä–∏–∞–Ω—Ç–∞–º–∏ –æ—Ç–≤–µ—Ç–æ–≤"""
    if not options_str:
        return []
    return [opt.strip() for opt in str(options_str).split(';') if opt.strip()]


def parse_correct(correct_str):
    """–ü–∞—Ä—Å–∏–Ω–≥ –ø—Ä–∞–≤–∏–ª—å–Ω—ã—Ö –æ—Ç–≤–µ—Ç–æ–≤"""
    if not correct_str:
        return []
    matches = re.findall(r'([–ê-–Ø–ÅA-Z]\))', str(correct_str))
    return matches


def get_alice_image_id(image_name):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ ID –∫–∞—Ä—Ç–∏–Ω–∫–∏ –¥–ª—è –ê–ª–∏—Å—ã"""
    if not image_name:
        return None
    return ALICE_IMAGE_IDS.get(str(image_name).strip())


# –ó–∞–≥—Ä—É–∑–∫–∞ –≤—Å–µ—Ö –≤–æ–ø—Ä–æ—Å–æ–≤ –∏–∑ Excel
quizzes = {}
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        question, options, correct, explanation, image = (row + (None, None, None, None, None))[:5]
        if not question:
            continue

        alice_image_id = get_alice_image_id(image)

        data.append({
            "–í–æ–ø—Ä–æ—Å": str(question).strip(),
            "–í–∞—Ä–∏–∞–Ω—Ç—ã": parse_options(options),
            "–ü—Ä–∞–≤–∏–ª—å–Ω—ã–π": parse_correct(correct),
            "–ü–æ—è—Å–Ω–µ–Ω–∏–µ": str(explanation).strip() if explanation else "",
            "–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ": alice_image_id
        })
    quizzes[sheet_name] = data


def get_random_question(topic, previous_questions=None):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å–ª—É—á–∞–π–Ω–æ–≥–æ –≤–æ–ø—Ä–æ—Å–∞ –ø–æ —Ç–µ–º–µ"""
    if topic not in quizzes or not quizzes[topic]:
        return None

    if previous_questions is None:
        previous_questions = []

    available_questions = [q for q in quizzes[topic] if q["–í–æ–ø—Ä–æ—Å"] not in previous_questions]

    if not available_questions:
        available_questions = quizzes[topic]

    return random.choice(available_questions)


def normalize_answer(user_answer):
    """–ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –æ—Ç–≤–µ—Ç–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    if not user_answer:
        return ""

    user_answer = user_answer.strip().lower()

    digit_to_letter = {"1": "–∞", "2": "–±", "3": "–≤", "4": "–≥", "5": "–¥", "6": "–µ"}
    if user_answer in digit_to_letter:
        return digit_to_letter[user_answer]

    user_answer = re.sub(r'[).\s,]', '', user_answer)

    if user_answer and user_answer[0] in '–∞–±–≤–≥–¥–µ':
        return user_answer[0]

    return ""


def normalize_correct_answers(correct_answers):
    """–ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –ø—Ä–∞–≤–∏–ª—å–Ω—ã—Ö –æ—Ç–≤–µ—Ç–æ–≤"""
    normalized = []
    for answer in correct_answers:
        clean_answer = re.sub(r'[)\s]', '', answer).lower()
        if clean_answer and clean_answer[0] in '–∞–±–≤–≥–¥–µ':
            normalized.append(clean_answer[0])
    return normalized


def parse_multiple_answers(command):
    """–ü–∞—Ä—Å–∏–Ω–≥ –Ω–µ—Å–∫–æ–ª—å–∫–∏—Ö –æ—Ç–≤–µ—Ç–æ–≤ –∏–∑ –∫–æ–º–∞–Ω–¥—ã"""
    cleaned = re.sub(r'[.,;]', ' ', command.lower())
    answers = cleaned.split()

    normalized_answers = []
    valid_answers = set()

    for answer in answers:
        normalized = normalize_answer(answer)
        if normalized and normalized not in valid_answers:
            normalized_answers.append(normalized)
            valid_answers.add(normalized)

    return normalized_answers


# –•—Ä–∞–Ω–∏–ª–∏—â–µ —Å–µ—Å—Å–∏–π –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
user_sessions = {}
user_stats = {}  # –û—Ç–¥–µ–ª—å–Ω–æ–µ —Ö—Ä–∞–Ω–∏–ª–∏—â–µ –¥–ª—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏


def init_user_stats(session_id):
    """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    user_stats[session_id] = {
        "total_answered": 0,
        "correct_answers": 0,
        "incorrect_answers": 0,
        "skipped_questions": 0,
        "current_topic": None
    }
    return user_stats[session_id]


def get_user_stats(session_id):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    if session_id not in user_stats:
        return init_user_stats(session_id)
    return user_stats[session_id]


def update_user_stats(session_id, result_type):
    """–û–±–Ω–æ–≤–ª–µ–Ω–∏–µ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    stats = get_user_stats(session_id)

    # –ì–∞—Ä–∞–Ω—Ç–∏—Ä—É–µ–º, —á—Ç–æ –≤—Å–µ –∫–ª—é—á–∏ —Å—É—â–µ—Å—Ç–≤—É—é—Ç
    if "total_answered" not in stats:
        stats["total_answered"] = 0
    if "correct_answers" not in stats:
        stats["correct_answers"] = 0
    if "incorrect_answers" not in stats:
        stats["incorrect_answers"] = 0
    if "skipped_questions" not in stats:
        stats["skipped_questions"] = 0
    if "current_topic" not in stats:
        stats["current_topic"] = None

    if result_type == "correct":
        stats["correct_answers"] += 1
        stats["total_answered"] += 1
    elif result_type == "incorrect":
        stats["incorrect_answers"] += 1
        stats["total_answered"] += 1
    elif result_type == "skipped":
        stats["skipped_questions"] += 1


def get_progress_text(session_id):
    """–§–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–∏–µ —Ç–µ–∫—Å—Ç–∞ —Å –ø—Ä–æ–≥—Ä–µ—Å—Å–æ–º"""
    stats = get_user_stats(session_id)

    # –ì–∞—Ä–∞–Ω—Ç–∏—Ä—É–µ–º, —á—Ç–æ –≤—Å–µ –∫–ª—é—á–∏ —Å—É—â–µ—Å—Ç–≤—É—é—Ç
    total_attempted = stats.get("total_answered", 0)
    correct = stats.get("correct_answers", 0)
    incorrect = stats.get("incorrect_answers", 0)
    skipped = stats.get("skipped_questions", 0)

    if total_attempted == 0:
        return "–í—ã –µ—â–µ –Ω–µ –æ—Ç–≤–µ—Ç–∏–ª–∏ –Ω–∏ –Ω–∞ –æ–¥–∏–Ω –≤–æ–ø—Ä–æ—Å."

    accuracy = (correct / total_attempted * 100) if total_attempted > 0 else 0

    progress_text = (
        f"üìä –í–∞—à –ø—Ä–æ–≥—Ä–µ—Å—Å:\n"
        f"‚Ä¢ –í—Å–µ–≥–æ —Ä–µ—à–µ–Ω–æ: {total_attempted} –≤–æ–ø—Ä–æ—Å–æ–≤\n"
        f"‚Ä¢ –ü—Ä–∞–≤–∏–ª—å–Ω—ã—Ö –æ—Ç–≤–µ—Ç–æ–≤: {correct}\n"
        f"‚Ä¢ –ù–µ–ø—Ä–∞–≤–∏–ª—å–Ω—ã—Ö –æ—Ç–≤–µ—Ç–æ–≤: {incorrect}\n"
        f"‚Ä¢ –ü—Ä–æ–ø—É—â–µ–Ω–æ –≤–æ–ø—Ä–æ—Å–æ–≤: {skipped}\n"
        f"‚Ä¢ –¢–æ—á–Ω–æ—Å—Ç—å: {accuracy:.1f}%"
    )

    return progress_text


@app.route("/", methods=["POST"])
def main():
    """–û—Å–Ω–æ–≤–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ –∑–∞–ø—Ä–æ—Å–æ–≤ –æ—Ç –ê–ª–∏—Å—ã"""
    try:
        req = request.json
        if not req:
            return jsonify_error("–ü—É—Å—Ç–æ–π –∑–∞–ø—Ä–æ—Å")

        command = req["request"]["command"].strip().lower()
        session = req.get("session", {})
        session_id = session.get("session_id")

        logger.info(f"–ó–∞–ø—Ä–æ—Å: –∫–æ–º–∞–Ω–¥–∞='{command}', session_id={session_id}")

        user_state = user_sessions.get(session_id, {})
        user_statistics = get_user_stats(session_id)

        response = {
            "version": req["version"],
            "session": req["session"],
            "response": {"end_session": False, "text": "", "buttons": []},
            "session_state": {}
        }

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–º–∞–Ω–¥—ã –ø—Ä–æ–≥—Ä–µ—Å—Å–∞
        if any(progress_cmd in command for progress_cmd in
               ["–ø—Ä–æ–≥—Ä–µ—Å—Å", "—Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞", "—Å—Ç–∞—Ç–∞", "—Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã", "—Å–∫–æ–ª—å–∫–æ"]):
            progress_text = get_progress_text(session_id)
            response["response"]["text"] = progress_text
            response["response"]["buttons"] = [
                {"title": "–ü—Ä–æ–¥–æ–ª–∂–∏—Ç—å"},
                {"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"}
            ]
            logger.info("–ü–æ–∫–∞–∑–∞–Ω–∞ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞")
            return jsonify(response)

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –Ω–æ–≤–æ–π —Å–µ—Å—Å–∏–∏
        if session.get("new", False):
            user_sessions[session_id] = {}
            init_user_stats(session_id)  # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –¥–ª—è –Ω–æ–≤–æ–π —Å–µ—Å—Å–∏–∏
            buttons = [{"title": name} for name in sheet_names] + [{"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}]
            response["response"]["text"] = "–ü—Ä–∏–≤–µ—Ç! –í—ã–±–µ—Ä–∏—Ç–µ —Ç–µ–º—É –¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è:"
            response["response"]["buttons"] = buttons
            logger.info("–ù–æ–≤–∞—è —Å–µ—Å—Å–∏—è")
            return jsonify(response)

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–º–∞–Ω–¥—ã –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –º–µ–Ω—é
        if any(nav_cmd in command for nav_cmd in ["–Ω–∞–∑–∞–¥", "–º–µ–Ω—é", "–≥–ª–∞–≤–Ω–∞—è", "–≤—ã—Ö–æ–¥"]):
            user_sessions[session_id] = {}
            buttons = [{"title": name} for name in sheet_names] + [{"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}]
            response["response"]["text"] = "–í—ã –≤–µ—Ä–Ω—É–ª–∏—Å—å –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é. –í—ã–±–µ—Ä–∏—Ç–µ —Ç–µ–º—É:"
            response["response"]["buttons"] = buttons
            logger.info("–í–æ–∑–≤—Ä–∞—Ç –≤ –º–µ–Ω—é")
            return jsonify(response)

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –ø—Ä–æ–ø—É—Å–∫–∞ –≤–æ–ø—Ä–æ—Å–∞
        if any(skip_cmd in command for skip_cmd in ["–ø—Ä–æ–ø—É—Å—Ç–∏—Ç—å", "—Å–ª–µ–¥—É—é—â–∏–π", "–¥–∞–ª—å—à–µ", "skip", "next"]):
            if user_state.get("mode") == "question" and user_state.get("topic"):
                update_user_stats(session_id, "skipped")
                topic = user_state["topic"]
                previous_questions = user_state.get("previous_questions", [])

                next_question = get_random_question(topic, previous_questions)
                if next_question:
                    options_text = "\n".join([f"{opt}" for opt in next_question["–í–∞—Ä–∏–∞–Ω—Ç—ã"]]) if next_question[
                        "–í–∞—Ä–∏–∞–Ω—Ç—ã"] else ""

                    if next_question["–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ"]:
                        response["response"]["card"] = {
                            "type": "BigImage",
                            "image_id": next_question["–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ"],
                            "title": f"–¢–µ–º–∞: {topic}",
                            "description": f"{next_question['–í–æ–ø—Ä–æ—Å']}\n\n{options_text}"
                        }
                        response["response"]["text"] = f"–í–æ–ø—Ä–æ—Å –ø—Ä–æ–ø—É—â–µ–Ω. –°–º–æ—Ç—Ä–∏—Ç–µ –∫–∞—Ä—Ç–∏–Ω–∫—É —Å –≤–æ–ø—Ä–æ—Å–æ–º –≤—ã—à–µ."
                    else:
                        response_text = f"–í–æ–ø—Ä–æ—Å –ø—Ä–æ–ø—É—â–µ–Ω.\n\n–¢–µ–º–∞: \"{topic}\"\n\n{next_question['–í–æ–ø—Ä–æ—Å']}\n\n{options_text}"
                        if len(response_text) > 1000:
                            response_text = response_text[:997] + "..."
                        response["response"]["text"] = response_text

                    updated_previous_questions = previous_questions + [next_question["–í–æ–ø—Ä–æ—Å"]]
                    user_sessions[session_id] = {
                        "topic": topic,
                        "question": next_question,
                        "previous_questions": updated_previous_questions,
                        "mode": "question"
                    }
                    user_statistics["current_topic"] = topic
                else:
                    response["response"]["text"] = "–í–æ–ø—Ä–æ—Å—ã –≤ —ç—Ç–æ–π —Ç–µ–º–µ –∑–∞–∫–æ–Ω—á–∏–ª–∏—Å—å."
                    user_sessions[session_id] = {}

                response["response"]["buttons"] = [
                    {"title": "–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å"},
                    {"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"},
                    {"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}
                ]
                logger.info("–í–æ–ø—Ä–æ—Å –ø—Ä–æ–ø—É—â–µ–Ω")
                return jsonify(response)

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–º–∞–Ω–¥—ã –ø–æ–º–æ—â–∏
        if command in ["–ø–æ–º–æ—â—å", "help", "—á—Ç–æ –¥–µ–ª–∞—Ç—å", "–ø—Ä–∞–≤–∏–ª–∞"]:
            if user_state.get("mode") == "question":
                response["response"][
                    "text"] = f"–í—ã –≤ —Ä–µ–∂–∏–º–µ –≤–æ–ø—Ä–æ—Å–∞ –ø–æ —Ç–µ–º–µ '{user_state['topic']}'. –ü—Ä–æ–∏–∑–Ω–µ—Å–∏—Ç–µ –Ω–æ–º–µ—Ä –æ—Ç–≤–µ—Ç–∞ (1-6) –∏–ª–∏ –±—É–∫–≤—É (–ê-–ï). –ú–æ–∂–Ω–æ –Ω–µ—Å–∫–æ–ª—å–∫–æ –æ—Ç–≤–µ—Ç–æ–≤ —á–µ—Ä–µ–∑ –ø—Ä–æ–±–µ–ª. –°–∫–∞–∂–∏—Ç–µ '–ø—Ä–æ–ø—É—Å—Ç–∏—Ç—å' –¥–ª—è –ø–µ—Ä–µ—Ö–æ–¥–∞ –∫ —Å–ª–µ–¥—É—é—â–µ–º—É –≤–æ–ø—Ä–æ—Å—É. –ò–ª–∏ —Å–∫–∞–∂–∏—Ç–µ '–Ω–∞–∑–∞–¥' –¥–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –º–µ–Ω—é. –¢–∞–∫–∂–µ –º–æ–∂–µ—Ç–µ –ø–æ—Å–º–æ—Ç—Ä–µ—Ç—å —Å–≤–æ–π –ø—Ä–æ–≥—Ä–µ—Å—Å, —Å–∫–∞–∑–∞–≤ '–ø—Ä–æ–≥—Ä–µ—Å—Å'."
            else:
                response["response"][
                    "text"] = "–Ø –ø–æ–º–æ–≥—É –≤–∞–º –ø–æ–¥–≥–æ—Ç–æ–≤–∏—Ç—å—Å—è –∫ —ç–∫–∑–∞–º–µ–Ω—É. –í—ã–±–µ—Ä–∏—Ç–µ —Ç–µ–º—É –¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –∏–ª–∏ —Å–∫–∞–∂–∏—Ç–µ '–Ω–∞–∑–∞–¥' –≤ –ª—é–±–æ–π –º–æ–º–µ–Ω—Ç. –í–æ –≤—Ä–µ–º—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –º–æ–∂–Ω–æ –ø—Ä–æ–ø—É—Å–∫–∞—Ç—å –≤–æ–ø—Ä–æ—Å—ã –∫–æ–º–∞–Ω–¥–æ–π '–ø—Ä–æ–ø—É—Å—Ç–∏—Ç—å'. –°–∫–∞–∂–∏—Ç–µ '–ø—Ä–æ–≥—Ä–µ—Å—Å' —á—Ç–æ–±—ã —É–≤–∏–¥–µ—Ç—å —Å–≤–æ—é —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É."
            response["response"]["buttons"] = [{"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"}, {"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}]
            logger.info("–ü–æ–∫–∞–∑–∞–Ω–∞ –ø–æ–º–æ—â—å")
            return jsonify(response)

        # –í—ã–±–æ—Ä —Ç–µ–º—ã
        for sheet_name in sheet_names:
            if (command == sheet_name.lower() or
                    (sheet_name == "–ü–µ—Ä–≤–∞—è –ø–æ–º–æ—â—å" and command in ["1 –ø–æ–º–æ—â—å", "–ø–µ—Ä–≤–∞—è –ø–æ–º–æ—â—å", "1–ø–æ–º–æ—â—å",
                                                                   "–ø–µ—Ä–≤–∞—è–ø–æ–º–æ—â—å"])):

                topic = sheet_name
                question = get_random_question(topic)
                if not question:
                    response["response"]["text"] = f"–í —Ç–µ–º–µ '{topic}' –Ω–µ—Ç –≤–æ–ø—Ä–æ—Å–æ–≤."
                    response["response"]["buttons"] = [{"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"}, {"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}]
                    logger.warning(f"–í —Ç–µ–º–µ '{topic}' –Ω–µ—Ç –≤–æ–ø—Ä–æ—Å–æ–≤")
                    return jsonify(response)

                options_text = "\n".join([f"{opt}" for opt in question["–í–∞—Ä–∏–∞–Ω—Ç—ã"]]) if question["–í–∞—Ä–∏–∞–Ω—Ç—ã"] else ""

                if question["–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ"]:
                    response["response"]["card"] = {
                        "type": "BigImage",
                        "image_id": question["–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ"],
                        "title": f"–¢–µ–º–∞: {topic}",
                        "description": f"{question['–í–æ–ø—Ä–æ—Å']}\n\n{options_text}"
                    }
                    response["response"]["text"] = f"–°–º–æ—Ç—Ä–∏—Ç–µ –≤–æ–ø—Ä–æ—Å –Ω–∞ –∫–∞—Ä—Ç–∏–Ω–∫–µ. {question['–í–æ–ø—Ä–æ—Å']}"
                else:
                    response_text = f'–¢–µ–º–∞: "{topic}"\n\n{question["–í–æ–ø—Ä–æ—Å"]}\n\n{options_text}'
                    if len(response_text) > 1000:
                        response_text = response_text[:997] + "..."
                    response["response"]["text"] = response_text

                response["response"]["buttons"] = [
                    {"title": "–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å"},
                    {"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"},
                    {"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}
                ]

                user_sessions[session_id] = {
                    "topic": topic,
                    "question": question,
                    "previous_questions": [question["–í–æ–ø—Ä–æ—Å"]],
                    "mode": "question"
                }
                user_statistics["current_topic"] = topic

                logger.info(f"–í—ã–±—Ä–∞–Ω–∞ —Ç–µ–º–∞ '{topic}'")
                return jsonify(response)

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –æ—Ç–≤–µ—Ç–∞ –Ω–∞ –≤–æ–ø—Ä–æ—Å
        if user_state.get("mode") == "question" and user_state.get("topic") and user_state.get("question"):
            topic = user_state["topic"]
            current_question = user_state["question"]
            previous_questions = user_state.get("previous_questions", [])

            logger.info(f"–û–±—Ä–∞–±–æ—Ç–∫–∞ –æ—Ç–≤–µ—Ç–∞ –¥–ª—è —Ç–µ–º—ã '{topic}': '{command}'")

            user_answers = parse_multiple_answers(command)
            correct_answers_normalized = normalize_correct_answers(current_question["–ü—Ä–∞–≤–∏–ª—å–Ω—ã–π"])

            logger.info(f"–û—Ç–≤–µ—Ç—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è: {user_answers}")
            logger.info(f"–ü—Ä–∞–≤–∏–ª—å–Ω—ã–µ –æ—Ç–≤–µ—Ç—ã: {correct_answers_normalized}")

            if not user_answers:
                response["response"][
                    "text"] = f"–ù–µ –ø–æ–Ω—è–ª –æ—Ç–≤–µ—Ç '{command}'. –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ —Ü–∏—Ñ—Ä—ã 1-6 –∏–ª–∏ –±—É–∫–≤—ã –ê-–ï. –ü—Ä–∏–º–µ—Ä: '1', '–∞', '1 2', '–∞ –±'. –°–∫–∞–∂–∏—Ç–µ '–ø—Ä–æ–ø—É—Å—Ç–∏—Ç—å' –¥–ª—è –ø–µ—Ä–µ—Ö–æ–¥–∞ –∫ —Å–ª–µ–¥—É—é—â–µ–º—É –≤–æ–ø—Ä–æ—Å—É. –ò–ª–∏ —Å–∫–∞–∂–∏—Ç–µ '–Ω–∞–∑–∞–¥' –¥–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –º–µ–Ω—é."
                response["response"]["buttons"] = [
                    {"title": "–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å"},
                    {"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"},
                    {"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}
                ]
                user_sessions[session_id] = user_state
                return jsonify(response)

            correct_given = [ans for ans in user_answers if ans in correct_answers_normalized]
            incorrect_given = [ans for ans in user_answers if ans not in correct_answers_normalized]

            # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞ –∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏
            if not incorrect_given and len(correct_given) == len(correct_answers_normalized):
                text = "–í–µ—Ä–Ω–æ!"
                update_user_stats(session_id, "correct")
            elif not incorrect_given and len(correct_given) > 0:
                missing = [ans for ans in correct_answers_normalized if ans not in user_answers]
                missing_text = ", ".join([f"{ans.upper()})" for ans in missing])
                text = f"–ß–∞—Å—Ç–∏—á–Ω–æ –≤–µ—Ä–Ω–æ! –í—ã –≤—ã–±—Ä–∞–ª–∏ –ø—Ä–∞–≤–∏–ª—å–Ω—ã–µ –æ—Ç–≤–µ—Ç—ã, –Ω–æ –Ω–µ —Ö–≤–∞—Ç–∞–µ—Ç: {missing_text}\n\n{current_question['–ü–æ—è—Å–Ω–µ–Ω–∏–µ']}"
                update_user_stats(session_id, "incorrect")
            elif len(correct_given) > 0 and len(incorrect_given) > 0:
                correct_text = ", ".join([f"{ans.upper()})" for ans in correct_given])
                incorrect_text = ", ".join([f"{ans.upper()})" for ans in incorrect_given])
                text = f"–ß–∞—Å—Ç–∏—á–Ω–æ –≤–µ—Ä–Ω–æ! –ü—Ä–∞–≤–∏–ª—å–Ω—ã–µ: {correct_text}, –Ω–µ–ø—Ä–∞–≤–∏–ª—å–Ω—ã–µ: {incorrect_text}\n\n{current_question['–ü–æ—è—Å–Ω–µ–Ω–∏–µ']}"
                update_user_stats(session_id, "incorrect")
            else:
                correct_text = ", ".join(current_question["–ü—Ä–∞–≤–∏–ª—å–Ω—ã–π"])
                text = f"–ù–µ–≤–µ—Ä–Ω–æ.\n–ü—Ä–∞–≤–∏–ª—å–Ω—ã–π –æ—Ç–≤–µ—Ç: {correct_text}\n\n{current_question['–ü–æ—è—Å–Ω–µ–Ω–∏–µ']}"
                update_user_stats(session_id, "incorrect")

            next_question = get_random_question(topic, previous_questions)
            if next_question:
                options_text = "\n".join([f"{opt}" for opt in next_question["–í–∞—Ä–∏–∞–Ω—Ç—ã"]]) if next_question[
                    "–í–∞—Ä–∏–∞–Ω—Ç—ã"] else ""

                if next_question["–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ"]:
                    response["response"]["card"] = {
                        "type": "BigImage",
                        "image_id": next_question["–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ"],
                        "title": f"–¢–µ–º–∞: {topic}",
                        "description": f"{next_question['–í–æ–ø—Ä–æ—Å']}\n\n{options_text}"
                    }
                    text += f"\n\n–°–ª–µ–¥—É—é—â–∏–π –≤–æ–ø—Ä–æ—Å: —Å–º–æ—Ç—Ä–∏—Ç–µ –Ω–∞ –∫–∞—Ä—Ç–∏–Ω–∫–µ –≤—ã—à–µ."
                else:
                    text += f"\n\n–°–ª–µ–¥—É—é—â–∏–π –≤–æ–ø—Ä–æ—Å:\n{next_question['–í–æ–ø—Ä–æ—Å']}\n\n{options_text}"

                if len(text) > 1000:
                    text = text[:997] + "..."

                updated_previous_questions = previous_questions + [next_question["–í–æ–ø—Ä–æ—Å"]]
                user_sessions[session_id] = {
                    "topic": topic,
                    "question": next_question,
                    "previous_questions": updated_previous_questions,
                    "mode": "question"
                }
                user_statistics["current_topic"] = topic
            else:
                text += "\n\n–í–æ–ø—Ä–æ—Å—ã –≤ —ç—Ç–æ–π —Ç–µ–º–µ –∑–∞–∫–æ–Ω—á–∏–ª–∏—Å—å."
                user_sessions[session_id] = {}

            response["response"]["text"] = text
            response["response"]["buttons"] = [
                {"title": "–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å"},
                {"title": "–ù–∞–∑–∞–¥ –≤ –º–µ–Ω—é"},
                {"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}
            ]
            return jsonify(response)

        # –ö–æ–º–∞–Ω–¥–∞ –Ω–µ —Ä–∞—Å–ø–æ–∑–Ω–∞–Ω–∞
        buttons = [{"title": name} for name in sheet_names] + [{"title": "–ü—Ä–æ–≥—Ä–µ—Å—Å"}]
        response["response"][
            "text"] = "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ —Ç–µ–º—É –∏–∑ –ø—Ä–µ–¥–ª–æ–∂–µ–Ω–Ω—ã—Ö –Ω–∏–∂–µ –∏–ª–∏ —Å–∫–∞–∂–∏—Ç–µ '–ø—Ä–æ–≥—Ä–µ—Å—Å' –¥–ª—è –ø—Ä–æ—Å–º–æ—Ç—Ä–∞ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏."
        response["response"]["buttons"] = buttons
        return jsonify(response)

    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∑–∞–ø—Ä–æ—Å–∞: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify_error("–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑.")


def jsonify_error(message):
    """–§–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–∏–µ –æ—Ç–≤–µ—Ç–∞ —Å –æ—à–∏–±–∫–æ–π"""
    return jsonify({
        "version": "1.0",
        "response": {"text": message, "end_session": False},
        "session_state": {}
    })


@app.route("/", methods=["GET"])
def home():
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ GET –∑–∞–ø—Ä–æ—Å–æ–≤ –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ä–∞–±–æ—Ç—ã —Å–µ—Ä–≤–µ—Ä–∞"""
    total_sessions = len(user_sessions)
    active_sessions = sum(1 for session in user_sessions.values() if session.get("mode") == "question")

    return jsonify({
        "status": "success",
        "message": "–ù–∞–≤—ã–∫ –ê–ª–∏—Å—ã —Ä–∞–±–æ—Ç–∞–µ—Ç.",
        "active_sessions": active_sessions,
        "total_sessions": total_sessions,
        "topics_loaded": list(quizzes.keys())
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info(f"–ó–∞–ø—É—Å–∫ —Å–µ—Ä–≤–µ—Ä–∞ –Ω–∞ –ø–æ—Ä—Ç—É {port}")
    app.run(host="0.0.0.0", port=port, debug=False)